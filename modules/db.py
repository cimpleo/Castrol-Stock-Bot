from openpyxl import load_workbook, Workbook
import logging

class DB:
    data = dict()

    def __init__(self, table_file):
        wb: Workbook = load_workbook(filename=table_file)
        sheets = wb.get_sheet_names()

        try:
            sheet = wb[sheets[0]]
            row_count = sheet.max_row

            for row in range(2, row_count + 1):
                sku = sheet.cell(column=1, row=row).value
                name = sheet.cell(column=2, row=row).value
                stock = sheet.cell(column=3, row=row).value
                count = sheet.cell(column=4, row=row).value
                record = dict(sku=sku, name=name, stock=stock, count=count)

                key: str = record['sku'].strip().upper()
                if key not in self.data:
                    self.data[key] = list()

                self.data[key].append(record)

        except IndexError:
            print('No sheets found in the data file')

        logging.info(f'Table loaded: {self.data}')

from openpyxl import load_workbook, Workbook


class DB:
    data = dict()

    def __init__(self, table_file):
        wb: Workbook = load_workbook(filename=table_file)
        sheet = wb['Sheet1']
        row_count = sheet.max_row

        for row in range(2, row_count):
            sku = sheet.cell(column=1, row=row).value
            name = sheet.cell(column=2, row=row).value
            stock = sheet.cell(column=3, row=row).value
            count = sheet.cell(column=4, row=row).value
            record = dict(sku=sku, name=name, stock=stock, count=count)
            self.data[record['sku']] = record

        print(self.data)